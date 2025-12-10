"""
Excel export pipeline.
Exports job listings to formatted Excel files with multiple sheets.
"""

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from src.models.job import JobListing
from src.utils.logging import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """
    Exports job listings to Excel with formatting.
    
    Creates separate sheets for:
    - All Jobs: All scraped jobs
    - Filtered Jobs: Jobs after filtering
    - With Cover Letters: Jobs with generated cover letters
    """
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    
    def __init__(self, output_path: str | Path):
        """
        Initialize the exporter.
        
        Args:
            output_path: Directory to save Excel files
        """
        self.output_path = Path(output_path)
        self.output_path.mkdir(parents=True, exist_ok=True)
    
    def _jobs_to_dataframe(self, jobs: list[JobListing]) -> pd.DataFrame:
        """
        Convert job listings to DataFrame.
        
        Args:
            jobs: List of job listings
        
        Returns:
            DataFrame with job data
        """
        if not jobs:
            return pd.DataFrame()
        
        data = [job.to_dict() for job in jobs]
        df = pd.DataFrame(data)
        
        # Reorder columns for better readability
        column_order = [
            "Title", "Budget", "Client Location", "Posted", 
            "Skills", "Description", "Job URL", "Cover Letter", "Scraped At"
        ]
        
        # Only include columns that exist
        columns = [c for c in column_order if c in df.columns]
        return df[columns]
    
    def _style_worksheet(self, ws, df: pd.DataFrame) -> None:
        """
        Apply styling to worksheet.
        
        Args:
            ws: Worksheet to style
            df: DataFrame that was written
        """
        if df.empty:
            return
        
        # Style header row
        for cell in ws[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.CELL_BORDER
        
        # Style data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.CELL_BORDER
        
        # Set column widths
        column_widths = {
            "A": 40,  # Title
            "B": 15,  # Budget
            "C": 20,  # Location
            "D": 15,  # Posted
            "E": 30,  # Skills
            "F": 60,  # Description
            "G": 50,  # URL
            "H": 60,  # Cover Letter
            "I": 20,  # Scraped At
        }
        
        for col, width in column_widths.items():
            if col <= chr(ord("A") + ws.max_column - 1):
                ws.column_dimensions[col].width = width
        
        # Set row height for better readability
        ws.row_dimensions[1].height = 30
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 60
    
    def export(
        self,
        all_jobs: list[JobListing],
        filtered_jobs: Optional[list[JobListing]] = None,
        jobs_with_letters: Optional[list[JobListing]] = None,
        filename: Optional[str] = None,
    ) -> Path:
        """
        Export jobs to Excel file.
        
        Args:
            all_jobs: All scraped jobs
            filtered_jobs: Jobs after filtering (optional)
            jobs_with_letters: Jobs with cover letters (optional)
            filename: Custom filename (optional)
        
        Returns:
            Path to the exported file
        """
        # Generate filename
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"upwork_jobs_{timestamp}.xlsx"
        
        filepath = self.output_path / filename
        
        logger.info(f"Exporting to: {filepath}")
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: All Jobs
        ws_all = wb.active
        ws_all.title = "All Jobs"
        df_all = self._jobs_to_dataframe(all_jobs)
        
        for row in dataframe_to_rows(df_all, index=False, header=True):
            ws_all.append(row)
        self._style_worksheet(ws_all, df_all)
        
        # Sheet 2: Filtered Jobs (if different from all)
        if filtered_jobs is not None and len(filtered_jobs) != len(all_jobs):
            ws_filtered = wb.create_sheet("Filtered Jobs")
            df_filtered = self._jobs_to_dataframe(filtered_jobs)
            
            for row in dataframe_to_rows(df_filtered, index=False, header=True):
                ws_filtered.append(row)
            self._style_worksheet(ws_filtered, df_filtered)
        
        # Sheet 3: Jobs with Cover Letters
        if jobs_with_letters:
            # Only include jobs that have cover letters
            jobs_with_cl = [j for j in jobs_with_letters if j.cover_letter]
            
            if jobs_with_cl:
                ws_letters = wb.create_sheet("With Cover Letters")
                df_letters = self._jobs_to_dataframe(jobs_with_cl)
                
                for row in dataframe_to_rows(df_letters, index=False, header=True):
                    ws_letters.append(row)
                self._style_worksheet(ws_letters, df_letters)
        
        # Add summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Jobs Scraped", len(all_jobs)])
        ws_summary.append(["Jobs After Filtering", len(filtered_jobs) if filtered_jobs else len(all_jobs)])
        ws_summary.append(["Jobs with Cover Letters", len([j for j in (jobs_with_letters or []) if j.cover_letter])])
        ws_summary.append(["Export Time", datetime.now().isoformat()])
        
        # Style summary
        for cell in ws_summary[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
        
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 30
        
        # Save
        wb.save(filepath)
        logger.info(f"Excel export complete: {filepath}")
        
        return filepath
    
    def export_cover_letters_txt(
        self,
        jobs: list[JobListing],
        output_dir: Optional[Path] = None,
    ) -> list[Path]:
        """
        Export cover letters as individual text files.
        
        Args:
            jobs: Jobs with cover letters
            output_dir: Directory to save files (default: output_path/cover_letters)
        
        Returns:
            List of paths to exported files
        """
        if output_dir is None:
            output_dir = self.output_path / "cover_letters"
        
        output_dir.mkdir(parents=True, exist_ok=True)
        
        exported = []
        
        for i, job in enumerate(jobs):
            if not job.cover_letter:
                continue
            
            # Create safe filename from job title
            safe_title = "".join(c if c.isalnum() or c in " -_" else "_" for c in job.title)
            safe_title = safe_title[:50]  # Limit length
            filename = f"{i+1:03d}_{safe_title}.txt"
            
            filepath = output_dir / filename
            
            content = f"""JOB: {job.title}
URL: {job.job_url}
BUDGET: {job.budget or job.hourly_rate or 'Not specified'}
SKILLS: {', '.join(job.skills)}

---

COVER LETTER:

{job.cover_letter}
"""
            
            filepath.write_text(content, encoding="utf-8")
            exported.append(filepath)
        
        logger.info(f"Exported {len(exported)} cover letters to {output_dir}")
        return exported
